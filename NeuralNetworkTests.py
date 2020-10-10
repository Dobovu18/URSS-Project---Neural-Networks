import ActivationFunctions as activation
import time
import numpy as np
import matplotlib.pyplot as plt
from ActivationFunctions import Sigmoid
import MultiLayerPerceptron as MLP
import openpyxl as xl

#DEFAULT PARAMETERS:
#input_nodes = 784
#hidden_nodes = 200
#hidden_layers = 2
#output_nodes = 10
#learning_rate = 1E-3
#training_file = "mnist_train.csv"
#test_file = "mnist_test.csv"
#sample_size = 20000
#interval = range(0, 100500, 500)
#activation_function = activation.Sigmoid
#activation_derivative = activation.d_Sigmoid

function_list = {
    "Sigmoid" : [activation.Sigmoid, activation.d_Sigmoid, '-'],
    "ReLU" : [activation.ReLU, activation.d_ReLU, '--'],
    "Leaky ReLU (a = 0.01)" : [activation.LeakyReLU, activation.d_LeakyReLU, '-.'],
    "Hyperbolic Tangent" : [activation.Tanh, activation.d_Tanh, ':'],
    "Arctan" : [activation.Arctan, activation.d_Arctan, '-'],
    "Identity" : [activation.Identity, activation.d_Identity, '--'],
    "Softplus" : [activation.Softplus, activation.d_Softplus, '-.'],
    "Softsign" : [activation.Softsign, activation.d_Softsign, ':'],
    "ELU (a = 0.01)" : [activation.ELU, activation.d_ELU, '-'],
    "ISRU (a = 0.01)" : [activation.ISRU, activation.d_ISRU, '--']
}

hiddenNodes_dict = {
    "50 nodes per layer"  : [50, '-'],
    "100 nodes per layer" : [100, '--'],
    "150 nodes per layer" : [150, '-.'],
    "200 nodes per layer" : [200, ':'],
    "250 nodes per layer" : [250, '-'],
    "300 nodes per layer" : [300, '--'],
    "350 nodes per layer" : [350, '-.'],
    "400 nodes per layer" : [400, ':'],
    "450 nodes per layer" : [450, '-'],
    "500 nodes per layer" : [500, '--'],
    "550 nodes per layer" : [550, '-.'],
    "600 nodes per layer" : [600, ':']
}

hiddenLayers_dict = {
    "1 hidden layer"   : [1, '-'],
    "2 hidden layers"  : [2, '--'],
    "3 hidden layers"  : [3, '-.'],
    "4 hidden layers"  : [4, ':'],
    "5 hidden layers"  : [5, '-'],
    "6 hidden layers"  : [6, '--'],
    "7 hidden layers"  : [7, '-.'],
    "8 hidden layers"  : [8, ':'],
    "9 hidden layers"  : [9, '-'],
    "10 hidden layers" : [10, '--']
}

learningRate_dict = {
    "Learning rate = 0.00" : [0.00, '-'],
    "Learning rate = 0.25" : [0.25, '--'],
    "Learning rate = 0.50" : [0.50, '-.'],
    "Learning rate = 0.75" : [0.75, ':'],
    "Learning rate = 1.00" : [1.00, '-'],
    "Learning rate = 1.25" : [1.25, '--'],
    "Learning rate = 1.50" : [1.50, '-.'],
    "Learning rate = 1.75" : [1.75, ':'],
    "Learning rate = 2.00" : [2.00, '-'],
    "Learning rate = 2.25" : [2.25, '--'],
    "Learning rate = 2.50" : [2.50, '-.'],
    "Learning rate = 2.75" : [2.75, ':'],
    "Learning rate = 3.00" : [3.00, '-']
}

start_time = time.time()

#Sigmoid_learningrates = np.linspace(0, 3, 31)
#learningrates = np.linspace(0, 3E-2, 31)
#learningrates = np.linspace(0, 1E-2, 26)


#Generating data of neural network performances
interval = range(0, 181000, 1000)

#Plotting data
#data = [[''] + list(interval)]
#for labels, var in hiddenLayers_dict.items(): #CHANGE
#    n = neuralNetwork(hiddenlayers = var[0], interval = interval) #CHANGE
#    n.train_from_file(3)
#    results = n.getPerformanceValues()
#    print(results)
#    data.append([labels] + results)
#    plt.plot(interval, results, label = labels, linestyle = var[1])
#

data = []
workbook = xl.Workbook()
sheet = workbook.active
counter = 0

Sigmoid_NN = MLP.neuralNetwork(
    activation_function = activation.Sigmoid,
    activation_derivative = activation.d_Sigmoid,
    learningrate = 0.8
)

ReLU_NN = MLP.neuralNetwork(
    activation_function = activation.ReLU,
    activation_derivative = activation.d_ReLU,
    learningrate = 7E-3
)

Tanh_NN = MLP.neuralNetwork(
    activation_function = activation.Tanh,
    activation_derivative=activation.d_Tanh,
    learningrate = 0.0084
)

LeakyReLU_NN = MLP.neuralNetwork(
    activation_function = activation.LeakyReLU,
    activation_derivative = activation.d_LeakyReLU,
    learningrate = 0.014
)

NN_dict = {
    "Sigmoid" : Sigmoid_NN,
    "ReLU" : ReLU_NN,
    "Tanh" : Tanh_NN,
    "LeakyReLU" : LeakyReLU_NN
}

for label, n in NN_dict.items():
    n.track_performance_and_train(epochs = 3, interval = interval)
    results = n.getPerformanceValues()
    record = [label + " -- learning rate =" + str(n.lr)] + results
    data.append(record)
    for i in range(len(interval)):
        sheet.cell(row = 1, column = i + 2, value = interval[i])
    for i in range(len(record)):
        sheet.cell(row = counter + 2, column = i + 1, value = record[i])
        workbook.save("final_comparison2.xlsx")  
    plt.plot(interval, results, label = label)
    counter += 1
print(data)  

#n = MLP.neuralNetwork(
#    learningrate=0.0096,
#    activation_function = activation.Tanh, 
#    activation_derivative=activation.d_Tanh,
#    sample_size=60000
#    )
#n.train_from_file(3)
#print(n.checkPerformance("mnist_test.csv"))
#for lr in learningrates:
#    n = MLP.neuralNetwork(
#        learningrate = lr,
#        activation_function = activation.LeakyReLU, #Change
#        activation_derivative = activation.d_LeakyReLU, #Change
#        sample_size = 60000
#    )
#    n.train_from_file(3)
#    result = n.checkPerformance("mnist_test.csv")
#    sheet.cell(row = 1, column = counter + 1, value = learningrates[counter])
#    sheet.cell(row = 2, column = counter + 1, value = result)
#    data.append(result)
#    workbook.save("LeakyReLU_LearningRatevsPerformance_1E-2.xlsx") #Change
#    counter += 1


print("Time: %.2fs" %(time.time() - start_time))
#Change
#plt.plot(learningrates, data, label = "Activation Function = LeakyReLU (a = 0.01)\nHidden Layers = 2\nNodes Per Hidden Layer = 200\nSample Size = 60000\nEpochs = 3")
plt.title("Performance vs Number Of Iterations")
plt.xlabel("Number Of Iterations")
plt.ylabel("Performance")
plt.grid(b=True,
        which='major',
        color='#666666', 
        linestyle='-')
plt.minorticks_on()
plt.grid(b=True,
        which='minor', 
        color='#999999', 
        linestyle='-', 
        alpha=0.2)
plt.legend()
plt.show()

#--[Tracking Neural Network Training Progress]---------------------------------------------------------------
#neuralNetwork_1 = neuralNetwork(784, 100, 1, 10, 0.8, "mnist_train.csv", 100)
#neuralNetwork_1.train_from_file(10)
#plt.plot(neuralNetwork_1.getTrainingInfo()[0], neuralNetwork_1.getTrainingInfo[1])

#neuralNetwork_2 = neuralNetwork(784, 600, 5, 10, 3.0, "mnist_train.csv", 20000)

#--[Neural Network Cost]-------------------------------------------------------------------------------------
#cost_avg = []
#cost_min_err = []
#cost_max_err = []

#for x in epoch:
#    cost = neuralNetworkCostFunctionTest(epochs = x) #change
    
#    avg = np.average(cost)
#    min_err = avg - np.min(cost)
#    max_err = np.max(cost) - avg

#    cost_avg.append(avg)
#    cost_min_err.append(min_err)
#    cost_max_err.append(max_err)

#cost_err = [cost_min_err, cost_max_err]

#--[Neural Network Performance]------------------------------------------------------------------------------
#performance_avg = []
#performance_min_err = []
#performance_max_err = []

#Processing values to plot on the graph
#for x in epoch: #Change
#        performance = neuralNetworkPerformanceTest(epochs = x) #Change

#        avg = np.average(performance) 
#        min_err = avg - np.min(performance)
#        max_err = np.max(performance) - avg

#        performance_avg.append(avg)
#        performance_min_err.append(min_err)
#        performance_max_err.append(max_err)

#performance_err = [performance_min_err, performance_max_err]

#--[Graph Creation]------------------------------------------------------------------------------------------
#legendLabel = "hidden layers = 2\nnodes per hidden layer = 100\nlearning rate = 0.5\ntraining set size = 20000" #change
#Plotting/Creating the graph
#plt.errorbar(epoch, cost_avg, #Change
#            yerr = cost_err,
#            label = legendLabel,
#            fmt = '-x',
#            ecolor = 'red',
#            capsize = 2)

#plt.title("Number Of Epochs vs Value Of Cost Function") #Change
#plt.xlabel("Number Of Epochs") #Change
#plt.ylabel("Value Of Cost Function")
#plt.legend()
#plt.grid(b=True,
#        which='major',
#        color='#666666', 
#        linestyle='-')
#plt.minorticks_on()
#plt.grid(b=True,
#        which='minor', 
#        color='#999999', 
#        linestyle='-', 
#        alpha=0.2)
#plt.show()